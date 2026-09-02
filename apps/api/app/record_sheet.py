"""SUFA 경기기록지(xlsx)에서 라인업을 뽑아낸다.

시트 한 장이 한 경기이고, 좌우로 홈/어웨이가 나뉜 고정 서식이다. 기존
기록지와 815대회용 분리형 템플릿을 모두 읽는다.

기존 기록지:
    홈    A=배번/위치  B=이름  C=교체(시간/이름)
    어웨이 K=배번/위치  L=이름  M=교체(시간/이름)

815 분리형 템플릿:
    홈    A=배번  B=포지션  C=이름  D=교체 정보
    어웨이 J=배번  K=포지션  L=이름  M=교체 정보
    선발 11명은 14~24행, 교체 10명은 26~35행에 둔다.

    팀명은 10행, 헤더는 13행, 선수는 14행부터 빈 줄까지.

실제 기록지에서 마주친 표기들을 그대로 처리한다.
    "45/GK"             배번/포지션
    "68 / GK"           슬래시 앞뒤 공백
    "11/LB 18"          원 등번호 / 포지션 + 조끼 번호(빨강, 맨 뒤)
    "99/RA.M", "6//LDM"  오타 — 점이 섞이거나 슬래시가 겹친다
    "허재원 ©", "장윤서c"  주장 표기 — 이름에서 떼어낸다
    "17'(66김시우)"       교체 (분, 들어온 선수 배번+이름)
    "50' (14김시원)"      괄호 앞 공백
    "50+3'(...)"         추가시간
    한 셀에 줄바꿈으로 교체 2건

조끼(bib) 번호
------------
조끼를 입고 뛴 팀은 등번호가 조끼 번호로 바뀐다. 기록지는 그 번호를 **빨간 글자**로
적는다. 실물(2026 SUFA 0405·0426 기록지 43건)은 전부 `원등번호/포지션 조끼번호` 꼴로,
조끼 번호가 포지션 **뒤에** 붙는다:

    "1/GK 1"  "11/LB 18"  "41/CDM 15"        ← 뒤 숫자가 빨강

영상에 보이는 건 조끼 번호이므로 그쪽을 배번으로 삼는다(원 번호는 rosterNumber 로
같이 실어 대조할 수 있게 남긴다). 조끼 번호가 어디에 붙든 상관없이 **색으로만** 고르므로
"7/LB"(조끼만)·"99/7/LB" 같은 다른 배치도 같은 코드로 처리된다.

색은 셀 전체 서식일 수도, 셀 안 일부 서식(rich text run)일 수도 있어 둘 다 본다.
그래서 워크북을 rich_text=True 로 연다 — 평범한 셀은 그대로 str 이라 기존 동작은
바뀌지 않는다.
"""
from __future__ import annotations

import io
import re
import unicodedata
from typing import Any

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock

TEAM_ROW = 10
FIRST_PLAYER_ROW = 14
# 기존 양식: 배번열, 이름열, 교체열, 팀명열
LEGACY_SIDES: dict[str, tuple[int, int, int, int]] = {
    "home": (1, 2, 3, 1),
    "away": (11, 12, 13, 11),
}
# 815 분리형 양식: 배번열, 포지션열, 이름열, 교체열, 팀명열, 감독명열
COLUMN_SIDES: dict[str, tuple[int, int, int, int, int, int]] = {
    "home": (1, 2, 3, 4, 1, 3),
    "away": (10, 11, 12, 13, 10, 12),
}
COLUMN_STARTER_ROWS = range(14, 25)
COLUMN_SUBSTITUTE_ROWS = range(26, 36)
SUB_RE = re.compile(r"(\d+(?:\+\d+)?)\s*'\s*\(\s*(\d+)\s*([^)]*?)\s*\)")
NUM_RE = re.compile(r"\d+")


def _is_red(rgb: Any) -> bool:
    """ARGB 문자열이 붉은 계열인가.

    기록지는 보통 순빨강(FFFF0000)을 쓰지만 손으로 고른 붉은 톤도 섞인다. 빨강만
    콕 집으면 놓치고, 아무 색이나 받으면 검정·회색까지 조끼로 오인한다 —
    빨강 성분이 충분히 크고 다른 채널을 확실히 앞설 때만 인정한다.

    openpyxl 은 테마 색일 때 rgb 자리에 str 이 아닌 값을 넣어 두므로 타입도 본다.
    """
    if not isinstance(rgb, str) or len(rgb) < 6:
        return False
    try:
        r, g, b = (int(rgb[-6:][i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return False
    return r >= 130 and r - max(g, b) >= 60


def _font_is_red(font: Any) -> bool:
    color = getattr(font, "color", None)
    return _is_red(getattr(color, "rgb", None)) if color is not None else False


def _number_tokens(cell: Any) -> list[tuple[str, bool]]:
    """배번 셀의 숫자 토큰을 (숫자, 빨간글자인가) 로 순서대로 돌려준다.

    포지션 라벨(GK·LCB…)에는 숫자가 없으므로 셀 전체에서 뽑아도 섞이지 않는다.
    """
    value = getattr(cell, "value", None)
    cell_red = _font_is_red(getattr(cell, "font", None))
    if isinstance(value, CellRichText):
        out: list[tuple[str, bool]] = []
        for part in value:
            red = _font_is_red(part.font) if isinstance(part, TextBlock) else cell_red
            out.extend((tok, red) for tok in NUM_RE.findall(str(part)))
        return out
    # 엑셀이 7 을 7.0 으로 돌려주면 숫자 토큰이 ["7", "0"] 로 갈라진다. 조끼 판정이
    # "빨간 숫자가 둘" 로 오인되므로 먼저 등번호 표기로 정규화한다.
    return [(tok, cell_red) for tok in NUM_RE.findall(_jersey_number(value))]


def _split_number_position(raw: str, tokens: list[tuple[str, bool]]) -> tuple[str, str, str, bool, bool]:
    """배번 셀 -> (쓸 배번, 포지션, 원 등번호, 조끼 판정이 애매한가, 조끼 번호를 썼는가).

    포지션은 슬래시로 나눈 조각 중 **글자가 든 마지막 것**에서 알파벳만 추린 것이다.
    알파벳만 남기는 게 핵심이다 —

        "11/LB 18"  뒤에 조끼 번호가 붙는다 (실제 기록지의 조끼 표기)
        "99/RA.M"   오타로 점이 섞인다      -> RAM
        "6//LDM"    슬래시가 겹친다        -> LDM
        "68 / GK"   슬래시 앞뒤 공백

    통째로 가져오면 "LB 18" 같은 게 나와 아는 포지션 목록에 없게 되고, 그 선수만
    조용히 사전 배치에서 빠진다(조끼 팀이 배치가 안 되던 원인이 이것이었다).

    조끼 번호는 빨간 숫자 하나일 때만 인정한다. 둘 이상이 빨갛다면 셀 전체가 붉게
    칠해진 것일 수 있어 어느 쪽이 조끼인지 알 수 없다 — 그때는 종전대로 첫 숫자를
    쓰고 애매함을 표시해 올려보낸다(조용히 틀린 번호를 고르는 것보다 낫다).
    """
    position = ""
    for part in reversed([p.strip() for p in raw.split("/") if p.strip()]):
        letters = "".join(re.findall(r"[A-Za-z]+", part))
        if letters:
            position = letters.upper()
            break

    plain = [tok for tok, red in tokens if not red]
    reds = [tok for tok, red in tokens if red]
    fallback = (plain or reds or [raw])[0]

    if len(reds) == 1:
        bib = reds[0]
        # 원 등번호가 조끼와 같은 숫자면(실제로 "1/GK 1" 이 있다) 따로 들고 다닐 게 없다.
        # 그래도 조끼를 입었다는 사실은 남긴다 — rosterNumber 유무로 판단하면 이 경우를 놓친다.
        roster = next((n for n in plain if n != bib), "")
        return bib, position, roster, False, True
    if len(reds) > 1:
        return fallback, position, "", True, False
    return fallback, position, "", False, False


def _nfc(value: Any) -> str:
    """macOS·엑셀을 오가며 자모 분해(NFD)된 한글이 섞여 들어온다. 비교 전에 합친다."""
    return unicodedata.normalize("NFC", str(value or "")).strip()


def _jersey_number(value: Any) -> str:
    """엑셀 숫자 셀의 정수형 소수 표기를 등번호로 안전하게 읽는다.

    엑셀이 `7`을 숫자로 저장한 경우 파일에 따라 openpyxl이 7.0으로 돌려줄 수
    있다. 숫자가 아닌 문자를 전부 지우면 `7.0`이 `70`이 되므로, 정수형
    소수점만 먼저 7로 정규화한다.
    """
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = _nfc(value)
    match = re.fullmatch(r"(\d+)\.0+", text)
    return match.group(1) if match else text


def _clean_name(raw: Any) -> tuple[str, bool]:
    """이름에서 주장 표기를 떼고 (이름, 주장여부)."""
    s = _nfc(raw)
    captain = bool(re.search(r"[©ⓒ]|(?<=[가-힣])c$|\(C\)$", s, re.I))
    s = re.sub(r"\s*[©ⓒ]\s*$", "", s)
    s = re.sub(r"(?<=[가-힣])c$", "", s)
    s = re.sub(r"\s*\(C\)\s*$", "", s, flags=re.I)
    return s.strip(), captain


def _parse_subs(raw: Any) -> list[dict]:
    out = []
    for m in SUB_RE.finditer(_nfc(raw)):
        minute, _, added = m.group(1).partition("+")
        out.append({
            "minute": int(minute),
            "added": int(added) if added else 0,
            "inNumber": m.group(2),
            "inName": _nfc(m.group(3)),
        })
    return out


# ── 포지션 유추 ──────────────────────────────────────────────────────────
# 조끼를 입고 뛴 팀은 배번 칸에 번호만 적고 포지션을 비워두는 일이 잦다. 라벨이 없으면
# 자리가 안 잡혀 사전 배치가 통째로 빠지므로, 팀명 옆 포메이션과 **기록지의 행 순서**로
# 자리를 메운다. 기록지는 GK 부터 최전방까지 위에서 아래로 적는 게 규칙이라 순서가 곧
# 줄 구성이다.
#
# 유추한 자리는 어차피 킥오프 형태의 출발점이고 태거가 끌어 옮긴다 — 정확한 좌표가
# 아니라 "누가 어느 줄에 있는지" 를 세워주는 게 목적이다.

# 줄 종류별로 인원수에 맞는 라벨. 프런트의 RECORD_SHEET_POSITION_GRID 에 있는 라벨만 쓴다
# (없는 라벨을 내보내면 그 선수만 조용히 배치에서 빠진다).
ROW_LABELS: dict[str, dict[int, list[str]]] = {
    "DEF": {
        2: ["LCB", "RCB"],
        3: ["LCB", "CB", "RCB"],
        4: ["LB", "LCB", "RCB", "RB"],
        5: ["LWB", "LCB", "CB", "RCB", "RWB"],
    },
    "DM": {1: ["DM"], 2: ["LDM", "RDM"], 3: ["LDM", "DM", "RDM"]},
    "CM": {
        1: ["CM"],
        2: ["LCM", "RCM"],
        3: ["LM", "CM", "RM"],
        4: ["LM", "LCM", "RCM", "RM"],
        5: ["LM", "LCM", "CM", "RCM", "RM"],
    },
    "AM": {1: ["CAM"], 2: ["LAM", "RAM"], 3: ["LAM", "CAM", "RAM"]},
    "ATT": {
        1: ["ST"],
        2: ["LS", "RS"],
        3: ["LW", "ST", "RW"],
        4: ["LW", "LS", "RS", "RW"],
        5: ["LW", "LS", "ST", "RS", "RW"],
    },
}
# 중원이 몇 줄이냐에 따라 깊이를 나눈다. 4-2-3-1 이면 [DM, AM] 이 되어 실제 기록지
# 표기(LDM/RDM → LAM/CAM/RAM)와 그대로 맞는다.
MID_DEPTHS: dict[int, list[str]] = {1: ["CM"], 2: ["DM", "AM"], 3: ["DM", "CM", "AM"]}
# 포메이션을 못 읽었을 때 아웃필드 인원수로 잡는 기본 줄 구성.
DEFAULT_LINES: dict[int, list[int]] = {
    10: [4, 4, 2], 9: [4, 3, 2], 8: [3, 3, 2], 7: [3, 2, 2], 6: [3, 2, 1], 5: [2, 2, 1],
}


def _lines_from_formation(formation: str) -> list[int]:
    """'4-2-3-1' · '4231' -> [4, 2, 3, 1].

    구분자 없이 붙여 쓴 표기도 받는다 — 기록지 팀명 칸에 그렇게 적힌 시트가 있다.
    붙여 쓴 건 아웃필드 10명이 맞아떨어질 때만 줄로 쪼갠다(우연한 숫자와 구분).
    """
    parts = [p for p in re.split(r"[^0-9]+", formation or "") if p]
    if len(parts) >= 2:
        return [int(p) for p in parts]
    if len(parts) == 1 and 3 <= len(parts[0]) <= 5:
        digits = [int(ch) for ch in parts[0]]
        if all(d >= 1 for d in digits) and sum(digits) == 10:
            return digits
    return []


def _split_team_formation(team: str) -> tuple[str, str]:
    """팀명 칸 -> (팀명, 포메이션).

    보통은 괄호 안에 적지만("한체대 태풍(4-2-3-1)"), 괄호 없이 뒤에 붙여 쓴 시트도
    있다("광운대학교 KWPE 4231"). 그대로 두면 팀명에 숫자가 섞여 작업 제목과 매칭이
    어긋나고 포메이션도 못 읽는다.

    괄호가 없을 때는 **아웃필드 10명이 맞아떨어질 때만** 포메이션으로 본다 — 팀명에
    들어간 연도 같은 숫자(FC 2002)를 포메이션으로 오인하지 않으려는 것이다.
    """
    fm = re.search(r"\(([^)]*\d[^)]*)\)\s*$", team)
    if fm:
        return team[: fm.start()].strip(), fm.group(1)
    tail = re.search(r"[\s]([\d][\d\s-]*)$", team)
    if tail and _lines_from_formation(tail.group(1)):
        return team[: tail.start()].strip(), tail.group(1).strip()
    return team, ""


def _slot_plan(lines: list[int]) -> list[str]:
    """줄 구성 -> GK 부터 최전방까지 늘어놓은 자리 라벨. 못 만들면 빈 목록."""
    if not lines or any(n <= 0 for n in lines):
        return []
    mids = lines[1:-1]
    depths = MID_DEPTHS.get(len(mids), [])
    if mids and not depths:
        return []                                   # 중원 4줄 이상 — 유추하지 않는다

    plan = ["GK"]
    used: set[str] = set()

    def take(kinds: list[str], count: int) -> list[str] | None:
        """이 줄에 쓸 라벨. 원래 깊이가 인원을 못 담으면 담을 수 있는 다른 줄로 옮긴다."""
        for kind in kinds:
            labels = ROW_LABELS.get(kind, {}).get(count)
            if labels and not (used & set(labels)):
                return labels
        return None

    picked = take(["DEF"], lines[0])
    if picked is None:
        return []
    plan += picked
    used |= set(picked)

    for depth, count in zip(depths, mids):
        # 원래 깊이를 먼저, 안 되면 폭이 넓은 CM 줄로 (4-1-4-1 의 4줄 같은 경우)
        picked = take([depth, "CM", "DM", "AM"], count)
        if picked is None:
            return []
        plan += picked
        used |= set(picked)

    if len(lines) >= 2:
        picked = take(["ATT"], lines[-1])
        if picked is None:
            return []
        plan += picked
    return plan


def _infer_positions(players: list[dict], formation: str) -> str:
    """포지션이 빈 선수에게 자리를 채워 넣는다. 어떻게 유추했는지를 돌려준다("" = 안 함).

    이미 적혀 있는 라벨은 건드리지 않고 그 자리를 '쓴 것' 으로 표시한다 — 일부만 비어
    있어도 남은 자리에서 순서대로 메워진다.
    """
    blanks = [p for p in players if not p.get("position")]
    if not blanks:
        return ""

    lines = _lines_from_formation(formation)
    source = "formation"
    if not lines:
        lines = DEFAULT_LINES.get(max(len(players) - 1, 0), [])
        source = "count"
    plan = _slot_plan(lines)
    # 계획한 자리 수와 명단 인원이 다르면 줄이 어긋난 것이다 — 엉뚱한 자리를 주느니 만다.
    if not plan or len(plan) != len(players):
        return ""

    taken = {str(p.get("position") or "").upper() for p in players if p.get("position")}
    free = [label for label in plan if label not in taken]
    for player in blanks:
        if not free:
            break
        player["position"] = free.pop(0)
        player["positionInferred"] = True
    return source


def _is_column_layout(ws) -> bool:
    """815대회 분리형 헤더를 명확히 구별해 기존 기록지는 그대로 보존한다."""
    expected = ("배번", "포지션", "선수명")
    home = tuple(_nfc(ws.cell(13, column).value).replace(" ", "") for column in (1, 2, 3))
    away = tuple(_nfc(ws.cell(13, column).value).replace(" ", "") for column in (10, 11, 12))
    return home == expected and away == expected


def _parse_legacy_side(ws, side: str) -> dict:
    col_no, col_name, col_sub, col_team = LEGACY_SIDES[side]
    # 괄호 없이 뒤에 붙여 쓴 포메이션("광운대학교 KWPE 4231")까지 읽는다.
    team, formation = _split_team_formation(_nfc(ws.cell(TEAM_ROW, col_team).value))

    players: list[dict] = []
    for r in range(FIRST_PLAYER_ROW, ws.max_row + 1):
        cell_no_obj = ws.cell(r, col_no)
        cell_no = cell_no_obj.value
        cell_nm = ws.cell(r, col_name).value
        if not cell_no and not cell_nm:
            if players:
                break                      # 명단 끝
            continue
        # 엑셀이 7 을 7.0 으로 돌려주는 시트가 있어 먼저 정규화한 뒤(_jersey_number),
        # 조끼 번호(빨간 글자)를 색으로 가려낸다.
        raw_no = _jersey_number(cell_no)
        number, position, roster_number, bib_ambiguous, bib_used = _split_number_position(
            raw_no, _number_tokens(cell_no_obj),
        )
        position = _nfc(position)
        name, captain = _clean_name(cell_nm)
        if not name or not number:
            continue
        entry = {
            "jerseyNumber": number,
            "name": name,
            "position": position,
            "captain": captain,
            "subs": _parse_subs(ws.cell(r, col_sub).value),
        }
        # 조끼를 입은 선수 — 배번은 조끼 번호이고, 원 등번호는 대조용으로 남긴다.
        if bib_used:
            entry["bib"] = True
            if roster_number:
                entry["rosterNumber"] = roster_number
        elif bib_ambiguous:
            # 빨간 숫자가 여럿이라 조끼를 특정 못 했다. 화면에서 눈으로 확인하라는 표시.
            entry["bibAmbiguous"] = True
        players.append(entry)

    # 포지션 칸이 빈 선수 메우기 — 조끼 팀 기록지에서 자주 비어 있다.
    inferred = _infer_positions(players, formation)
    out = {"team": team, "formation": formation, "coach": "", "players": players}
    if inferred:
        out["positionsInferred"] = inferred
    return out


def _parse_column_side(ws, side: str) -> dict:
    col_no, col_position, col_name, col_sub, col_team, col_coach = COLUMN_SIDES[side]
    team, formation = _split_team_formation(_nfc(ws.cell(TEAM_ROW, col_team).value))
    coach = _nfc(ws.cell(37, col_coach).value)
    players: list[dict] = []

    for r in (*COLUMN_STARTER_ROWS, *COLUMN_SUBSTITUTE_ROWS):
        raw_no = _jersey_number(ws.cell(r, col_no).value)
        raw_position = _nfc(ws.cell(r, col_position).value).upper()
        raw_name = ws.cell(r, col_name).value
        if not raw_no and not raw_position and not _nfc(raw_name):
            # 선발·교체 사이 25행은 표의 구분 라벨이라 그냥 통과한다.
            continue
        name, captain = _clean_name(raw_name)
        if not raw_no or not name:
            continue
        players.append({
            "jerseyNumber": raw_no,
            "name": name,
            "position": raw_position,
            "captain": captain,
            "isSubstitute": r in COLUMN_SUBSTITUTE_ROWS,
            "subs": _parse_subs(ws.cell(r, col_sub).value),
        })
    return {"team": team, "formation": formation, "coach": coach, "players": players}


def parse_sheet(ws) -> dict:
    is_column_layout = _is_column_layout(ws)
    return {
        "matchNo": _nfc(ws.cell(7, 2).value),
        "date": _nfc(ws.cell(7, 5).value),
        "venue": _nfc(ws.cell(8, 5).value),
        "home": _parse_column_side(ws, "home") if is_column_layout else _parse_legacy_side(ws, "home"),
        "away": _parse_column_side(ws, "away") if is_column_layout else _parse_legacy_side(ws, "away"),
    }


def parse_workbook(file_bytes: bytes) -> list[dict]:
    """기록지 전체를 시트 목록으로 돌려준다. 서식이 다른 시트는 건너뛰지 않고 표시한다."""
    try:
        # rich_text=True — 조끼 번호는 셀 안 일부만 빨간 경우가 있어 run 단위 색이 필요하다.
        # 색이 없는 평범한 셀은 그대로 str 이라 기존 파싱 결과는 바뀌지 않는다.
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True, read_only=False, rich_text=True,
        )
    except Exception as exc:  # noqa: BLE001 - 업로드 파일이라 원인이 다양하다
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    out: list[dict] = []
    for name in wb.sheetnames:
        # 815 템플릿의 첫 탭은 작성 예시만 담은 가이드다. 실제 경기 시트로
        # 오인해 일괄 등록 목록에 표시하지 않는다.
        if _nfc(name).startswith("작성 가이드"):
            continue
        try:
            parsed = parse_sheet(wb[name])
        except Exception as exc:  # noqa: BLE001
            out.append({"sheet": name, "error": str(exc)[:200]})
            continue
        parsed["sheet"] = name
        parsed["usable"] = bool(
            parsed["home"]["players"] and parsed["away"]["players"]
        )
        out.append(parsed)
    return out
