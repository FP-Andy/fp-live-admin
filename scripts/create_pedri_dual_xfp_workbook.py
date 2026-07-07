#!/usr/bin/env python3
"""Generate a Pedri dual-pitch FPA workbook with a draft xFP sheet.

This is a product trace demo:
- FPA dual pitch state is encoded through generate_log_entry().
- The normal FPA export workbook is created through build_analysis_workbook().
- A Pedri_xFP sheet is appended with the current draft scoring rules.
"""

from __future__ import annotations

import json
import math
import sys
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "apps" / "api"))

from app.fpa import build_analysis_workbook, generate_log_entry, parse_logs_to_dataframe  # noqa: E402


REACTION_TIME = 0.7
SHARED_PLAYER_SPEED = 5.5
PC_TAU = 1.15

REF_CURVES = {
    "Goal": {"ref50": 0.068, "ref_hi": 0.188},
    "Creation": {"ref50": 0.068, "ref_hi": 0.188},
    "Progression": {"ref50": 0.0013, "ref_hi": 0.0083},
    "Possession": {"ref50": 0.030, "ref_hi": 0.120},
}

ACTION_CATALOG = {
    "A01": ("Shot/OPP/Direct/Goal", "Goal"),
    "A02": ("Pass/OPP/Indirect/Goal", "Creation"),
    "A05": ("Pass/OPP/Indirect/Progression", "Progression"),
    "A08": ("Dribble/OPP/Direct/Progression", "Progression"),
    "A09": ("Penetration/OPP/Direct/Progression", "Progression"),
    "A12": ("Pass/OPP/Direct/Possession", "Possession"),
    "A17": ("Defense/OPP/Direct/Possession", "Possession"),
}

MF_ROLE_WEIGHTS = {
    "Advanced Playmaker": {"A01": 2, "A02": 5, "A05": 4, "A08": 2, "A12": 4},
    "Box To Box": {"A01": 3, "A02": 2, "A05": 3, "A08": 3, "A09": 2, "A12": 3, "A17": 4},
    "Deep-Lying Playmaker": {"A05": 3, "A12": 3},
    "Mezzala": {"A01": 3, "A02": 4, "A05": 4, "A08": 4, "A09": 3, "A12": 3, "A17": 2},
}


@dataclass(frozen=True)
class PedriEvent:
    stat_input: str
    timeline: str
    dots: list[dict[str, float]]
    before: list[dict[str, Any]]
    after: list[dict[str, Any]]
    action_id: str
    metric_name: str
    base_value: float
    note: str
    packing_bypassed: int = 0
    dcm: float = 0.0
    credit: float = 1.0

    @property
    def action_key(self) -> str:
        return ACTION_CATALOG[self.action_id][0]

    @property
    def outcome(self) -> str:
        return ACTION_CATALOG[self.action_id][1]

    def raw_xfp(self, metric_value: float | None = None) -> float:
        base_value = self.base_value if metric_value is None else metric_value
        if self.outcome in {"Goal", "Creation"}:
            return base_value * self.credit
        packing_factor = min(1.6, 1 + 0.15 * max(0, self.packing_bypassed))
        return base_value * packing_factor * (1 + self.dcm) * self.credit


def dot(meter_x: float, meter_y: float, team_side: str, number: str = "", role: str = "field") -> dict[str, Any]:
    return {
        "meter_x": meter_x,
        "meter_y": meter_y,
        "team": "ally" if team_side == "home" else "opponent",
        "team_side": team_side,
        "role": role,
        "layer": f"{team_side}_{'gk' if role == 'gk' else 'field'}",
        "number": number,
        "id": f"{team_side}-{number or role}-{meter_x:.1f}-{meter_y:.1f}",
    }


def dual(before: list[dict[str, Any]], after: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "actor_team": "home",
        "primary_row_index": 0,
        "input_tier": "recommended",
        "before": {"dots": before},
        "after": {"dots": after},
    }


def split_team_points(points: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    home = [point for point in points if point.get("team_side") == "home"]
    away = [point for point in points if point.get("team_side") == "away"]
    return home, away


def team_control_at(target: dict[str, float], points: list[dict[str, Any]]) -> float:
    total = 0.0
    for point in points:
        distance = math.hypot(float(point["meter_x"]) - target["meter_x"], float(point["meter_y"]) - target["meter_y"])
        intercept_time = REACTION_TIME + distance / SHARED_PLAYER_SPEED
        total += math.exp(-intercept_time / PC_TAU)
    return total


def pitch_control_at(target: dict[str, float], points: list[dict[str, Any]]) -> float:
    home, away = split_team_points(points)
    home_control = team_control_at(target, home)
    away_control = team_control_at(target, away)
    denominator = home_control + away_control
    home_probability = home_control / denominator if denominator else 0.5
    return round((home_probability * 2) - 1, 4)


def pc_components(event: PedriEvent) -> dict[str, float]:
    source = event.dots[0]
    target = event.dots[-1]
    before_pc = pitch_control_at(source, event.before)
    after_pc = pitch_control_at(target, event.after)
    return {
        "before_pc": before_pc,
        "after_pc": after_pc,
        "pc_delta": round(max(0.0, after_pc - before_pc), 4),
    }


def nearest_points(target: dict[str, float], points: list[dict[str, Any]], team_side: str, limit: int = 2) -> str:
    candidates = [point for point in points if point.get("team_side") == team_side]
    ranked = sorted(
        candidates,
        key=lambda point: math.hypot(float(point["meter_x"]) - target["meter_x"], float(point["meter_y"]) - target["meter_y"]),
    )
    labels = []
    for point in ranked[:limit]:
        distance = math.hypot(float(point["meter_x"]) - target["meter_x"], float(point["meter_y"]) - target["meter_y"])
        labels.append(f"{point.get('number') or '?'}@({point['meter_x']},{point['meter_y']}) d={distance:.1f}")
    return "; ".join(labels)


def metric_value(event: PedriEvent) -> float:
    if event.metric_name == "pc_delta":
        return pc_components(event)["pc_delta"]
    return event.base_value


def add_metrics(log_text: str, metrics: dict[str, float]) -> str:
    if not metrics:
        return log_text
    parts = log_text.split(" | ")
    metric_text = ", ".join(f"{key}={value:.4f}" for key, value in metrics.items())
    metric_index = next((idx for idx, part in enumerate(parts) if part.startswith("Metrics: ")), -1)
    if metric_index >= 0:
        existing = parts[metric_index].replace("Metrics: ", "")
        parts[metric_index] = f"Metrics: {existing}, {metric_text}" if existing else f"Metrics: {metric_text}"
        return " | ".join(parts)
    dual_index = next((idx for idx, part in enumerate(parts) if part.startswith("DualState: ")), len(parts))
    parts.insert(dual_index, f"Metrics: {metric_text}")
    return " | ".join(parts)


def absolute_score(value: float, outcome: str) -> float:
    ref = REF_CURVES[outcome]
    k = math.log(85 / 15) / (ref["ref_hi"] - ref["ref50"])
    return 100 / (1 + math.exp(-k * (value - ref["ref50"])))


def pedri_events() -> list[PedriEvent]:
    base_before = [
        dot(44, 37, "home", "8"),
        dot(55, 45, "home", "10"),
        dot(40, 28, "home", "6"),
        dot(73, 34, "home", "9"),
        dot(50, 39, "away", "4"),
        dot(58, 33, "away", "6"),
        dot(68, 40, "away", "5"),
        dot(101, 34, "away", "1", "gk"),
    ]
    return [
        PedriEvent(
            stat_input="8ss10.k.p.sw",
            timeline="12:14",
            dots=[{"meter_x": 44, "meter_y": 37}, {"meter_x": 70, "meter_y": 43}],
            before=base_before,
            after=[
                dot(46, 38, "home", "8"),
                dot(70, 43, "home", "10"),
                dot(43, 29, "home", "6"),
                dot(78, 35, "home", "9"),
                dot(52, 39, "away", "4"),
                dot(61, 34, "away", "6"),
                dot(72, 41, "away", "5"),
                dot(101, 34, "away", "1", "gk"),
            ],
            action_id="A02",
            metric_name="linked_xg",
            base_value=0.110,
            note="Key pass through the midfield line into a shot-enabling zone.",
            packing_bypassed=2,
        ),
        PedriEvent(
            stat_input="8rr.p",
            timeline="18:03",
            dots=[{"meter_x": 48, "meter_y": 40}, {"meter_x": 62, "meter_y": 43}],
            before=base_before,
            after=[
                dot(62, 43, "home", "8"),
                dot(69, 45, "home", "10"),
                dot(44, 28, "home", "6"),
                dot(76, 34, "home", "9"),
                dot(53, 39, "away", "4"),
                dot(60, 36, "away", "6"),
                dot(72, 41, "away", "5"),
                dot(101, 34, "away", "1", "gk"),
            ],
            action_id="A08",
            metric_name="epv_delta",
            base_value=0.0049,
            note="Carry between lines with two defenders bypassed.",
            packing_bypassed=2,
            dcm=0.08,
        ),
        PedriEvent(
            stat_input="8pn10.p",
            timeline="23:41",
            dots=[{"meter_x": 57, "meter_y": 31}, {"meter_x": 76, "meter_y": 28}],
            before=base_before,
            after=[
                dot(76, 28, "home", "8"),
                dot(74, 42, "home", "10"),
                dot(46, 27, "home", "6"),
                dot(82, 34, "home", "9"),
                dot(58, 31, "away", "4"),
                dot(66, 30, "away", "6"),
                dot(78, 37, "away", "5"),
                dot(101, 34, "away", "1", "gk"),
            ],
            action_id="A09",
            metric_name="epv_delta",
            base_value=0.0055,
            note="Off-ball penetration run that opens a progressive receive lane.",
            packing_bypassed=1,
            dcm=0.05,
        ),
        PedriEvent(
            stat_input="8ss6.ret",
            timeline="39:08",
            dots=[{"meter_x": 64, "meter_y": 36}, {"meter_x": 58, "meter_y": 30}],
            before=base_before,
            after=[
                dot(64, 36, "home", "8"),
                dot(59, 30, "home", "6"),
                dot(70, 44, "home", "10"),
                dot(78, 34, "home", "9"),
                dot(61, 37, "away", "4"),
                dot(65, 31, "away", "6"),
                dot(75, 39, "away", "5"),
                dot(101, 34, "away", "1", "gk"),
            ],
            action_id="A12",
            metric_name="pc_delta",
            base_value=0.0,
            note="Tight-space pass retained in the opponent half.",
            dcm=0.05,
        ),
        PedriEvent(
            stat_input="8q",
            timeline="51:22",
            dots=[{"meter_x": 67, "meter_y": 38}],
            before=base_before,
            after=[
                dot(67, 38, "home", "8"),
                dot(72, 43, "home", "10"),
                dot(50, 28, "home", "6"),
                dot(80, 34, "home", "9"),
                dot(63, 38, "away", "4"),
                dot(69, 33, "away", "6"),
                dot(78, 40, "away", "5"),
                dot(101, 34, "away", "1", "gk"),
            ],
            action_id="A17",
            metric_name="pc_delta",
            base_value=0.0,
            note="Counterpress interception in the opponent half.",
            dcm=0.12,
        ),
        PedriEvent(
            stat_input="8dd.f",
            timeline="63:10",
            dots=[{"meter_x": 83, "meter_y": 38}],
            before=base_before,
            after=[
                dot(83, 38, "home", "8"),
                dot(75, 43, "home", "10"),
                dot(61, 29, "home", "6"),
                dot(88, 34, "home", "9"),
                dot(76, 37, "away", "4"),
                dot(80, 33, "away", "6"),
                dot(85, 40, "away", "5"),
                dot(101, 34, "away", "1", "gk"),
            ],
            action_id="A01",
            metric_name="xg",
            base_value=0.052,
            note="Low-volume zone-14 shot on target.",
            credit=0.95,
        ),
    ]


def build_logs(events: list[PedriEvent]) -> list[str]:
    logs: list[str] = []
    for event in events:
        result = generate_log_entry(
            stat_input=event.stat_input,
            dots=event.dots,
            half="1H" if event.timeline < "45:00" else "2H",
            team="home",
            direction="right",
            timeline=event.timeline,
            dual_pitch=dual(event.before, event.after),
        )
        metric_key = {"linked_xg": "xG", "xg": "xG", "epv_delta": "EPV", "pc_delta": "PC"}[event.metric_name]
        logs.append(add_metrics(result["log_text"], {metric_key: metric_value(event)}))
    return logs


def write_import_workbook(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)
        pd.DataFrame(
            [
                {
                    "Usage": "FPA > Live Logger > 수정 및 불러오기",
                    "Note": "이 파일은 Data 시트 중심의 FPA 재불러오기용 원본 로그입니다.",
                },
                {
                    "Usage": "Dual mode",
                    "Note": "DualState v0.2에 home/away 선수 좌표, GK, 등번호가 포함되어 수정용 피치에서 복원됩니다.",
                },
            ]
        ).to_excel(writer, sheet_name="Readme", index=False)


def append_pedri_xfp_sheet(workbook_bytes: bytes, events: list[PedriEvent], output_path: Path) -> None:
    wb = load_workbook(BytesIO(workbook_bytes))
    if "Pedri_xFP" in wb.sheetnames:
        del wb["Pedri_xFP"]
    ws = wb.create_sheet("Pedri_xFP", 0)

    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A1"] = "Pedri xFP draft sheet from FPA dual pitch mode"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:L1")
    ws["A2"] = "Persona"
    ws["B2"] = "Pedri"
    ws["C2"] = "Position group"
    ws["D2"] = "MF"
    ws["E2"] = "Status"
    ws["F2"] = "Dual state capture works; PC uses equal-speed home/away coordinate dominance; EPV remains draft/proxy."

    headers = [
        "No",
        "Time",
        "StatInput",
        "ActionID",
        "ActionKey",
        "Outcome",
        "Metric",
        "BaseValue",
        "Before_PC",
        "After_PC",
        "Packing",
        "DCM",
        "Credit",
        "Raw_xFP",
        "AbsScore",
        "DualEvidence",
    ]
    ws.append([])
    ws.append(headers)
    header_row = 4
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    action_scores: dict[str, list[float]] = {}
    for idx, event in enumerate(events, start=1):
        value = metric_value(event)
        pc = pc_components(event) if event.metric_name == "pc_delta" else {"before_pc": "", "after_pc": ""}
        raw = event.raw_xfp(value)
        score = absolute_score(raw, event.outcome)
        action_scores.setdefault(event.action_id, []).append(score)
        target = event.dots[-1]
        nearest_home = nearest_points(target, event.after, "home")
        nearest_away = nearest_points(target, event.after, "away")
        ws.append(
            [
                idx,
                event.timeline,
                event.stat_input,
                event.action_id,
                event.action_key,
                event.outcome,
                event.metric_name,
                value,
                pc["before_pc"],
                pc["after_pc"],
                event.packing_bypassed,
                event.dcm,
                event.credit,
                raw,
                score,
                f"{event.note} | nearest_home_after={nearest_home} | nearest_away_after={nearest_away}",
            ]
        )

    start_role_row = ws.max_row + 3
    ws.cell(start_role_row, 1, "Role ranking").font = Font(bold=True)
    ws.cell(start_role_row + 1, 1, "Role").font = Font(bold=True)
    ws.cell(start_role_row + 1, 2, "WeightedRoleScore").font = Font(bold=True)
    ws.cell(start_role_row + 1, 3, "ObservedActionCoverage").font = Font(bold=True)

    role_rows = []
    for role, weights in MF_ROLE_WEIGHTS.items():
        weighted = 0.0
        total = 0.0
        observed = 0
        for action_id, weight in weights.items():
            score = sum(action_scores[action_id]) / len(action_scores[action_id]) if action_id in action_scores else 50.0
            observed += int(action_id in action_scores)
            weighted += score * weight
            total += weight
        role_rows.append((role, weighted / total if total else 0.0, f"{observed}/{len(weights)}"))
    for role, score, coverage in sorted(role_rows, key=lambda row: row[1], reverse=True):
        ws.append([role, score, coverage])

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {
        "A": 8,
        "B": 10,
        "C": 14,
        "D": 10,
        "E": 34,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 10,
        "J": 10,
        "K": 9,
        "L": 8,
        "M": 8,
        "N": 12,
        "O": 12,
        "P": 72,
    }.items():
        ws.column_dimensions[col].width = width

    pc_ws = wb.create_sheet("Pedri_PC_Model", 1)
    pc_ws.append(["Pitch Control Model"])
    pc_ws.append(["Scale", "1 = home 100%, 0 = balanced, -1 = away 100%"])
    pc_ws.append(["Formula", "PC(z)=2*P_home(z)-1; P_home=sum(exp(-(reaction_time+distance/shared_speed)/tau) home)/sum(all)"])
    pc_ws.append(["Parameters", f"reaction_time={REACTION_TIME}, shared_player_speed={SHARED_PLAYER_SPEED}, tau={PC_TAU}"])
    pc_ws.append([])
    pc_headers = ["Time", "ActionID", "TargetX", "TargetY", "Before_PC", "After_PC", "PC_Delta", "NearestHomeAfter", "NearestAwayAfter"]
    pc_ws.append(pc_headers)
    for cell in pc_ws[6]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for event in events:
        target = event.dots[-1]
        pc = pc_components(event)
        pc_ws.append(
            [
                event.timeline,
                event.action_id,
                target["meter_x"],
                target["meter_y"],
                pc["before_pc"],
                pc["after_pc"],
                pc["pc_delta"],
                nearest_points(target, event.after, "home"),
                nearest_points(target, event.after, "away"),
            ]
        )
    for col, width in {"A": 10, "B": 10, "C": 9, "D": 9, "E": 10, "F": 10, "G": 10, "H": 44, "I": 44}.items():
        pc_ws.column_dimensions[col].width = width
    for row in pc_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    events = pedri_events()
    logs = build_logs(events)
    df = parse_logs_to_dataframe(logs, "pedri-dual-demo", "Pedri XI", "Opponent XI")
    runtime_dir = ROOT / "runtime" / "xfp_pedri_dual_demo"
    import_path = runtime_dir / "pedri_dual_mode_fpa_import.xlsx"
    write_import_workbook(df, import_path)
    workbook = build_analysis_workbook(df)
    output_path = runtime_dir / "pedri_dual_mode_xfp_demo.xlsx"
    append_pedri_xfp_sheet(workbook, events, output_path)

    log_path = output_path.with_suffix(".logs.json")
    log_path.write_text(json.dumps({"logs": logs}, indent=2), encoding="utf-8")
    print(import_path)
    print(output_path)
    print(log_path)


if __name__ == "__main__":
    main()
